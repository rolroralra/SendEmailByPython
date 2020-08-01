from smtpClient import SmtpClient
from openpyxl import load_workbook

import configparser
import os

if __name__ == '__main__':
    config = configparser.ConfigParser()
    config.read('config.ini')

    # SMTP Client
    smtpClient = SmtpClient(config['SMTP-SERVER'])

    # Email Send List
    emailSendListInfo = config['EMAIL-SEND-LIST']
    emailSendListFilePath = os.path.join(
        emailSendListInfo['FILE_DIR'],
        emailSendListInfo['FILE_NAME']
    )

    # Excel
    wb = load_workbook(emailSendListFilePath)
    ws = wb[emailSendListInfo['SHEET_NAME']]
    rows = ws[emailSendListInfo['START_COLUMN']:emailSendListInfo['END_COLUMN']]
    for row in ws.iter_rows():
        addr = row[0].value
        subj_layout = row[1].value
        cont_layout = row[2].value
        attachment = row[3].value

        # Send Mail
        print('RECEIVER:', addr)
        print('SUBJECT:', subj_layout)
        print('CONTENT:', cont_layout)
        print('ATTCHMENT:', attachment)
        smtpClient.send_mail(addr, subj_layout, cont_layout, attachment)
        print()
